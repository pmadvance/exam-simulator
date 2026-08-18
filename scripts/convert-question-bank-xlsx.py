#!/usr/bin/env python3
from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SOURCE_DIR = Path("docs/questions-sample/PMP_PRACTICE_EXAM_QUESTION")
OUT_DIR = SOURCE_DIR / "converted"

OUTPUT_HEADER = [
    "questionType",
    "prompt",
    "optionA",
    "optionB",
    "optionC",
    "optionD",
    "optionE",
    "correctAnswer",
    "explanation",
    "ecoDomain",
    "performanceDomain",
    "imageUrl",
    "status",
    "difficulty",
]


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("\n", " ")
    text = re.sub(r"[^a-z0-9]+", "", text)
    return text


HEADER_ALIASES = {
    "stem": "prompt",
    "stemandquestion": "prompt",
    "prompt": "prompt",
    "question": "prompt",
    "optiona": "optionA",
    "optionb": "optionB",
    "optionc": "optionC",
    "optiond": "optionD",
    "optione": "optionE",
    "key": "correctAnswer",
    "correctanswer": "correctAnswer",
    "feedback": "explanation",
    "feedbackandrationale": "explanation",
    "explanation": "explanation",
    "ecodomaintask": "ecoDomain",
    "ecomatch": "ecoDomain",
    "ecodomain": "ecoDomain",
    "classificationaagilehhybridppredictiveagagnostic": "performanceDomain",
    "classificationaagilehhybridppredictivexagnostic": "performanceDomain",
    "classification": "performanceDomain",
    "questiontype": "questionType",
    "type": "questionType",
    "status": "status",
    "difficulty": "difficulty",
    "imageurl": "imageUrl",
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    replacements = {
        '"': "'",
        "\u202f": " ",
        "\xa0": " ",
        "\t": " ",
        "\r\n": "\n",
        "\r": "\n",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = re.sub(r"[ ]+\n", "\n", text)
    text = re.sub(r"\n[ ]+", "\n", text)
    text = re.sub(r"[ ]{2,}", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def normalize_answer(value: Any) -> str:
    text = clean_text(value).upper()
    parts = re.findall(r"[A-E]", text)
    return ",".join(dict.fromkeys(parts))


def infer_question_type(answer: str, option_c: str, option_d: str, option_e: str) -> str:
    if len([part for part in answer.split(",") if part]) > 1:
        return "multiple_response"
    if not option_c and not option_d and not option_e:
        return "true_false"
    return "single_choice"


def normalize_classification(value: str) -> str:
    text = clean_text(value)
    compact = re.sub(r"[^A-Za-z]+", "", text).upper()
    lookup = {
        "A": "A - Agile",
        "A - Agile": "A - Agile",
        "A – Agile": "A - Agile",
        "H": "H - Hybrid",
        "H - Hybrid": "H - Hybrid",
        "H – Hybrid": "H - Hybrid",
        "P": "P - Predictive",
        "P - Predictive": "P - Predictive",
        "AG": "AG - Agnostic",
        "X": "AG - Agnostic",
        "X - Agnostic": "AG - Agnostic",
    }
    compact_lookup = {
        "AAGILE": "A - Agile",
        "HHYBRID": "H - Hybrid",
        "PPREDICTIVE": "P - Predictive",
        "AGAGNOSTIC": "AG - Agnostic",
        "XAGNOSTIC": "AG - Agnostic",
    }
    return lookup.get(text, compact_lookup.get(compact, text))


def make_index_map(header: tuple[Any, ...]) -> dict[str, int]:
    index_map: dict[str, int] = {}
    for index, value in enumerate(header):
        normalized = normalize_header(value)
        canonical = HEADER_ALIASES.get(normalized)
        if canonical and canonical not in index_map:
            index_map[canonical] = index
    return index_map


def get_cell(row: tuple[Any, ...], index_map: dict[str, int], key: str) -> str:
    index = index_map.get(key)
    if index is None or index >= len(row):
        return ""
    return clean_text(row[index])


def convert_workbook(path: Path) -> tuple[Path, int, int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    rows = sheet.iter_rows(values_only=True)
    header = next(rows)
    index_map = make_index_map(header)

    required = ["prompt", "optionA", "optionB", "correctAnswer", "explanation"]
    missing = [key for key in required if key not in index_map]
    if missing:
        raise RuntimeError(f"{path.name}: missing required columns: {', '.join(missing)}")

    output_rows: list[dict[str, str]] = []
    skipped = 0

    for row in rows:
        prompt = get_cell(row, index_map, "prompt")
        option_a = get_cell(row, index_map, "optionA")
        option_b = get_cell(row, index_map, "optionB")
        option_c = get_cell(row, index_map, "optionC")
        option_d = get_cell(row, index_map, "optionD")
        option_e = get_cell(row, index_map, "optionE")
        answer = normalize_answer(get_cell(row, index_map, "correctAnswer"))
        explanation = get_cell(row, index_map, "explanation")

        if not any([prompt, option_a, option_b, option_c, option_d, option_e, answer, explanation]):
            skipped += 1
            continue
        if not prompt or not option_a or not option_b or not answer or not explanation:
            skipped += 1
            continue

        question_type = get_cell(row, index_map, "questionType")
        if question_type:
            question_type = question_type.lower().replace(" ", "_")
        else:
            question_type = infer_question_type(answer, option_c, option_d, option_e)

        output_rows.append(
            {
                "questionType": question_type,
                "prompt": prompt,
                "optionA": option_a,
                "optionB": option_b,
                "optionC": "" if question_type == "true_false" else option_c,
                "optionD": "" if question_type == "true_false" else option_d,
                "optionE": "" if question_type == "true_false" else option_e,
                "correctAnswer": answer,
                "explanation": explanation,
                "ecoDomain": get_cell(row, index_map, "ecoDomain"),
                "performanceDomain": normalize_classification(get_cell(row, index_map, "performanceDomain")),
                "imageUrl": get_cell(row, index_map, "imageUrl"),
                "status": "published",
                "difficulty": get_cell(row, index_map, "difficulty"),
            }
        )

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUT_DIR / f"{path.stem}_import.csv"
    with out_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=OUTPUT_HEADER)
        writer.writeheader()
        writer.writerows(output_rows)

    return out_path, len(output_rows), skipped


def main() -> None:
    for path in sorted(SOURCE_DIR.glob("*.xlsx")):
        out_path, written, skipped = convert_workbook(path)
        print(f"{path.name} -> {out_path.name}: {written} rows, skipped {skipped}")


if __name__ == "__main__":
    main()
